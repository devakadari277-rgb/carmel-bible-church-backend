import random
import os
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from rest_framework import viewsets, permissions, status, generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate

from .models import User, Profile, ChurchSetting, PrayerRequest, Event, Announcement, LiveStream, Gallery, ContactMessage, ActivityLog, UserNotification
from .serializers import (
    UserSerializer, UserRegisterSerializer, ChurchSettingSerializer,
    PrayerRequestSerializer, EventSerializer, AnnouncementSerializer,
    LiveStreamSerializer, GallerySerializer, ContactMessageSerializer,
    ActivityLogSerializer, ProfileSerializer, UserNotificationSerializer
)

# Helper to generate JWT tokens
def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    # Add role to token claims
    refresh['role'] = user.role
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }

# Custom Permission classes
class IsAdminOrReadOnly(permissions.BasePermission):
    """
    Allows read-only access for members/anonymous users, 
    but write actions only for admins.
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user.is_authenticated and request.user.role == 'admin'

class IsAdminUserOnly(permissions.BasePermission):
    """
    Allows access only to authenticated admin users.
    """
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.role == 'admin'

class IsOwnerOrAdmin(permissions.BasePermission):
    """
    Allows write/delete access only to the owner of the object or an admin.
    """
    def has_object_permission(self, request, view, obj):
        # Admin can do anything
        if request.user and request.user.role == 'admin':
            return True
        # Owner can update/delete their own object
        return obj.user == request.user

# --- Authentication APIs ---

class RegisterView(generics.CreateAPIView):
    serializer_class = UserRegisterSerializer
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            tokens = get_tokens_for_user(user)
            user_data = UserSerializer(user).data
            return Response({
                'user': user_data,
                'tokens': tokens
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class LoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        email_or_username = request.data.get('email_or_username')
        password = request.data.get('password')

        if not email_or_username or not password:
            return Response({'error': 'Please provide email/username and password.'}, status=status.HTTP_400_BAD_REQUEST)

        # Authenticate user
        user = None
        if '@' in email_or_username:
            try:
                user_obj = User.objects.get(email__iexact=email_or_username)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
        else:
            user = authenticate(request, username=email_or_username, password=password)

        if not user:
            return Response({'error': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)

        # Return standard JWT tokens immediately
        tokens = get_tokens_for_user(user)
        user_data = UserSerializer(user).data
        return Response({
            'status': 'success',
            'user': user_data,
            'tokens': tokens
        }, status=status.HTTP_200_OK)

class AdminLoginView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request, *args, **kwargs):
        email_or_username = request.data.get('email') or request.data.get('email_or_username')
        password = request.data.get('password')

        if not email_or_username or not password:
            return Response({'error': 'Please provide email/username and password.'}, status=status.HTTP_400_BAD_REQUEST)

        user = None
        if '@' in email_or_username:
            try:
                user_obj = User.objects.get(email__iexact=email_or_username)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                pass
        else:
            user = authenticate(request, username=email_or_username, password=password)

        if not user:
            return Response({'error': 'Invalid credentials.'}, status=status.HTTP_401_UNAUTHORIZED)

        # Check if the user is an admin
        admin_emails = [e.strip().lower() for e in getattr(settings, 'ADMIN_EMAILS', [])]
        user_email = user.email.strip().lower() if user.email else ''

        if user_email not in admin_emails and user.role != 'admin':
            return Response({'error': 'Access Denied. You do not have administrator permissions.'}, status=status.HTTP_403_FORBIDDEN)

        # Generate tokens directly
        tokens = get_tokens_for_user(user)
        user_data = UserSerializer(user).data

        # Log admin login action
        ActivityLog.objects.create(
            user=user,
            action="Logged in successfully to Admin Portal."
        )

        return Response({
            'status': 'success',
            'user': user_data,
            'tokens': tokens
        }, status=status.HTTP_200_OK)


# --- Dashboard & Profile APIs ---

class ProfileView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response(serializer.data)

    def put(self, request):
        serializer = UserSerializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class AdminStatsView(APIView):
    permission_classes = [IsAdminUserOnly]

    def get(self, request):
        total_members = User.objects.filter(role='member').count()
        total_events = Event.objects.count()
        total_prayers = PrayerRequest.objects.count()
        total_live_videos = LiveStream.objects.count()
        recent_messages = ContactMessage.objects.order_by('-created_at')[:5]
        recent_logs = ActivityLog.objects.order_by('-created_at')[:10]

        message_serializer = ContactMessageSerializer(recent_messages, many=True)
        log_serializer = ActivityLogSerializer(recent_logs, many=True)

        return Response({
            'total_members': total_members,
            'total_events': total_events,
            'total_prayers': total_prayers,
            'total_live_videos': total_live_videos,
            'recent_messages': message_serializer.data,
            'recent_activities': log_serializer.data
        })


# --- Member & Content CRUD ViewSets ---

class UserViewSet(viewsets.ModelViewSet):
    """
    CRUD ViewSet for user list. Only Admin can list/delete members.
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'destroy', 'export_excel']:
            return [IsAdminUserOnly()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        # Exclude admins from standard members list
        return User.objects.filter(role='member').order_by('-date_joined')

    def perform_destroy(self, instance):
        action_msg = f"Deleted user account: {instance.username} ({instance.email})"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUserOnly])
    def export_excel(self, request):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registered Members"

        headers = ["Full Name", "Username", "Email", "Phone Number", "Registration Date"]
        
        # Apply style
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        users = User.objects.filter(role='member').order_by('-date_joined')
        for user in users:
            phone = user.profile.phone_number if hasattr(user, 'profile') else ""
            reg_date = user.date_joined.strftime("%Y-%m-%d %H:%M:%S") if user.date_joined else ""
            row = [
                user.first_name, # Name
                user.username,
                user.email,
                phone,
                reg_date
            ]
            ws.append(row)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 1:
                    cell.alignment = align_left if cell.column != 5 else align_center
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="registered_members.xlsx"'
        wb.save(response)
        return response

class ChurchSettingViewSet(viewsets.ModelViewSet):
    queryset = ChurchSetting.objects.all()
    serializer_class = ChurchSettingSerializer
    permission_classes = [IsAdminOrReadOnly]

    def list(self, request, *args, **kwargs):
        # Returns the single singleton ChurchSetting instance
        instance = ChurchSetting.get_settings()
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    def perform_update(self, serializer):
        serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action="Updated church settings and homepage information."
        )

class EventViewSet(viewsets.ModelViewSet):
    queryset = Event.objects.all().order_by('event_date')
    serializer_class = EventSerializer
    permission_classes = [IsAdminOrReadOnly]

    def perform_create(self, serializer):
        event = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Created upcoming event: '{event.title}'"
        )
        # Generate user notifications for all registered members
        members = User.objects.filter(role='member')
        for m in members:
            UserNotification.objects.create(
                user=m,
                title="New Event Published",
                message=f"Admin published a new upcoming event: '{event.title}' at {event.location}."
            )

    def perform_update(self, serializer):
        event = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Updated event details: '{event.title}'"
        )

    def perform_destroy(self, instance):
        action_msg = f"Deleted event: '{instance.title}'"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

    @action(detail=True, methods=['post'], permission_classes=[permissions.AllowAny])
    def notify_members(self, request, pk=None):
        event = self.get_object()
        if event.notified:
            return Response({'status': 'already_notified'}, status=status.HTTP_200_OK)

        # Get all registered members
        members = User.objects.filter(role='member')
        recipient_list = [m.email for m in members if m.email]

        if recipient_list:
            subject = f"Carmel Bible Church Fellowship - {event.title}"
            message = (
                f"Dear Church Member,\n\n"
                f"Our fellowship event '{event.title}' is starting now!\n\n"
                f"Details:\n"
                f"Date & Time: {event.event_date}\n"
                f"Location: {event.location}\n"
                f"Description: {event.description}\n\n"
                f"We hope to see you there!\n\n"
                f"Blessings,\n"
                f"Carmel Bible Church"
            )
            try:
                send_mail(
                    subject=subject,
                    message=message,
                    from_email="no-reply@carmelbiblechurch.org",
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
            except Exception as e:
                pass

        event.notified = True
        event.save()
        return Response({'status': 'notified', 'sent_to': len(recipient_list)}, status=status.HTTP_200_OK)

class PrayerRequestViewSet(viewsets.ModelViewSet):
    queryset = PrayerRequest.objects.all().order_by('-created_at')
    serializer_class = PrayerRequestSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.AllowAny()]
        if self.action in ['create', 'my_requests']:
            return [permissions.IsAuthenticated()]
        if self.action in ['update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsOwnerOrAdmin()]
        return [IsAdminUserOnly()]

    def get_queryset(self):
        user = self.request.user
        if user.is_authenticated and user.role == 'admin':
            return PrayerRequest.objects.all().order_by('-created_at')
        # Standard members can only see approved requests that have visibility set to 'all'
        return PrayerRequest.objects.filter(status='approved', visibility='all').order_by('-created_at')

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """
        Endpoint to retrieve requests filed by the logged-in member
        """
        queryset = PrayerRequest.objects.filter(user=request.user).order_by('-created_at')
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer):
        # Public requests ('all') are approved immediately. Private requests ('admin') start as pending.
        visibility = self.request.data.get('visibility', 'all')
        status_val = 'approved' if visibility == 'all' else 'pending'
        serializer.save(user=self.request.user, status=status_val)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        prayer = self.get_object()
        prayer.status = 'approved'
        prayer.save()
        ActivityLog.objects.create(
            user=request.user,
            action=f"Approved prayer request: '{prayer.title}' by {prayer.user.username}"
        )
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        prayer = self.get_object()
        prayer.status = 'rejected'
        prayer.save()
        ActivityLog.objects.create(
            user=request.user,
            action=f"Rejected prayer request: '{prayer.title}' by {prayer.user.username}"
        )
        return Response({'status': 'rejected'})

    @action(detail=True, methods=['post'])
    def toggle_pin(self, request, pk=None):
        prayer = self.get_object()
        prayer.is_pinned = not prayer.is_pinned
        prayer.save()
        action_name = "Pinned" if prayer.is_pinned else "Unpinned"
        ActivityLog.objects.create(
            user=request.user,
            action=f"{action_name} prayer request: '{prayer.title}'"
        )
        return Response({'status': 'success', 'is_pinned': prayer.is_pinned})

    def perform_destroy(self, instance):
        action_msg = f"Deleted prayer request: '{instance.title}' by {instance.user.username}"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

class AnnouncementViewSet(viewsets.ModelViewSet):
    queryset = Announcement.objects.all().order_by('-created_at')
    serializer_class = AnnouncementSerializer
    permission_classes = [IsAdminOrReadOnly]

    def perform_create(self, serializer):
        ann = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Published announcement: '{ann.title}'"
        )
        # Generate user notifications for all registered members
        members = User.objects.filter(role='member')
        for m in members:
            UserNotification.objects.create(
                user=m,
                title="New Announcement Posted",
                message=f"Admin posted a new announcement: '{ann.title}'."
            )

    def perform_update(self, serializer):
        ann = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Updated announcement details: '{ann.title}'"
        )

    def perform_destroy(self, instance):
        action_msg = f"Deleted announcement: '{instance.title}'"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

class LiveStreamViewSet(viewsets.ModelViewSet):
    queryset = LiveStream.objects.all().order_by('-created_at')
    serializer_class = LiveStreamSerializer
    permission_classes = [IsAdminOrReadOnly]

    def perform_create(self, serializer):
        stream = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Added video link: '{stream.title}'"
        )
        # Generate user notifications for all registered members
        members = User.objects.filter(role='member')
        for m in members:
            UserNotification.objects.create(
                user=m,
                title="New Ceremony Video Added",
                message=f"Admin uploaded a new ceremony/video: '{stream.title}'."
            )

    def perform_update(self, serializer):
        stream = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Updated video details: '{stream.title}'"
        )

    def perform_destroy(self, instance):
        action_msg = f"Deleted video: '{instance.title}'"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

class GalleryViewSet(viewsets.ModelViewSet):
    queryset = Gallery.objects.all().order_by('-uploaded_at')
    serializer_class = GallerySerializer
    permission_classes = [IsAdminOrReadOnly]

    def perform_create(self, serializer):
        photo = serializer.save()
        ActivityLog.objects.create(
            user=self.request.user,
            action=f"Uploaded new image to gallery: '{photo.caption}'"
        )
        # Generate user notifications for all registered members
        members = User.objects.filter(role='member')
        for m in members:
            UserNotification.objects.create(
                user=m,
                title="New Photo Gallery Upload",
                message=f"Admin added a new photo to the gallery: '{photo.caption}'."
            )

    def perform_destroy(self, instance):
        action_msg = f"Deleted image from gallery: '{instance.caption}'"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

class ContactMessageViewSet(viewsets.ModelViewSet):
    queryset = ContactMessage.objects.all().order_by('-created_at')
    serializer_class = ContactMessageSerializer

    def get_permissions(self):
        if self.action in ['create']:
            return [permissions.AllowAny()]
        return [IsAdminUserOnly()]

    def perform_destroy(self, instance):
        action_msg = f"Deleted contact message from: '{instance.name}'"
        ActivityLog.objects.create(user=self.request.user, action=action_msg)
        instance.delete()

class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ActivityLog.objects.all().order_by('-created_at')
    serializer_class = ActivityLogSerializer
    permission_classes = [IsAdminUserOnly]


class UserNotificationViewSet(viewsets.ModelViewSet):
    serializer_class = UserNotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return UserNotification.objects.filter(user=self.request.user).order_by('-created_at')

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        UserNotification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'status': 'marked_all_read'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'marked_read'}, status=status.HTTP_200_OK)
